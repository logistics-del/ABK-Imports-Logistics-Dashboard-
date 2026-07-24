"""Excel export helpers for the state-level pending-cases download."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXPORT_COLUMNS = [
    "Invoice Number", "Month", "Ship Date", "EDD Date", "Delivery Date",
    "Shipment Status", "Ageing TAT (days)", "City", "State",
    "Customer Mobile Number", "Mode of Payment", "Remarks", "Compliance Category",
]

# Statuses considered "pending / active" for the state click-through export.
PENDING_ACTIVE_STATUSES = ["INTRANSIT", "OFD", "EXCEPTION", "NDR", "MISS"]

COMPLIANCE_LABELS = {
    "CUSTOMER_ISSUE": "Customer Issue",
    "VENDOR_ISSUE": "Vendor Issue",
    "OTHER_ISSUE": "Other Issue",
}


def build_pending_cases_workbook(shipments, state_name):
    """Builds an in-memory .xlsx workbook of pending/active shipments for a
    given state, with exactly the columns required by the spec."""
    wb = Workbook()
    ws = wb.active
    ws.title = state_name[:31] if state_name else "Pending Cases"

    header_fill = PatternFill(start_color="13233F", end_color="13233F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, shipment in enumerate(shipments, start=2):
        ws.cell(row=row_idx, column=1, value=shipment.invoice_number)
        ws.cell(row=row_idx, column=2, value=shipment.month)
        ws.cell(row=row_idx, column=3, value=shipment.ship_date.isoformat() if shipment.ship_date else "")
        ws.cell(row=row_idx, column=4, value=shipment.edd_date.isoformat() if shipment.edd_date else "")
        ws.cell(row=row_idx, column=5, value=shipment.delivery_date.isoformat() if shipment.delivery_date else "")
        ws.cell(row=row_idx, column=6, value=shipment.shipment_status)
        ws.cell(row=row_idx, column=7, value=shipment.ageing_tat)
        ws.cell(row=row_idx, column=8, value=shipment.city)
        ws.cell(row=row_idx, column=9, value=shipment.state)
        ws.cell(row=row_idx, column=10, value=shipment.customer_mobile)
        ws.cell(row=row_idx, column=11, value=shipment.mode_of_payment)
        ws.cell(row=row_idx, column=12, value=shipment.remarks)
        ws.cell(row=row_idx, column=13, value=COMPLIANCE_LABELS.get(shipment.compliance, shipment.compliance))

    for col_idx, header in enumerate(EXPORT_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(14, len(header) + 4)

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
